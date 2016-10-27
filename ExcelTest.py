from openpyxl import Workbook, load_workbook
# wb = Workbook()

wb = load_workbook('document.xlsx')


# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")